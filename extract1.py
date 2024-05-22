from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import re
import glob2
# -------- Mọi thứ đều có công sức từ mọi người ở FOE-BAV, hoàn thiện và phát triển nó thật tốt -----
# Contact trao đổi : dangnguyen110900@gmail.com
ten_file = r'MAP OFP 12052023-09052024.xlsx'

def copy_text(start_pattern, end_pattern, text):
    start_match = re.search(start_pattern, text)
    end_match = re.search(end_pattern, text)

    if start_match and end_match:
        start_index = start_match.end()
        end_index = end_match.start()

        copied_text = text[start_index:end_index]
        return copied_text

def generate_data(txt_files):
    data = []
    i = 0
    for i, file_path in enumerate(txt_files, start=1):
        # Read the text file and extract its content
        with open(file_path, 'r') as file:
            text = file.read()
            start_pattern = r"- - -"
            end_pattern = r"END OF JEPPESEN DATAPLAN"
            copied_text = copy_text(start_pattern, end_pattern, text)

            row_data = [i, "", "", 
                        '= IFERROR(MID(@[CFP],FIND("TRIP TIME",@[CFP],1)-36,3)," ")', '= IFERROR(MID(@[CFP],FIND("TRIP TIME",@[CFP],1)-32,3)," ")',
                        "", '= IFERROR(MID(@[CFP],FIND("TRIP TIME",@[CFP],1)-6,3)," ")', '=IFERROR(MID(@[CFP],FIND("COMPANY ROUTE",@[CFP],1)+14,12)," ")', 
                        '=IFERROR(MID(@[CFP],FIND("THANG",@[CFP],1)+5,2)," ")', 
                        '=TEXT(IFERROR(DATE("20"&RIGHT(IFERROR(MID(@[CFP],FIND("FIFR  ",@[CFP],1)+6,8)," "),2),MID(IFERROR(MID(@[CFP],FIND("FIFR  ",@[CFP],1)+6,8)," "),4,2),LEFT(IFERROR(MID(@[CFP],FIND("FIFR  ",@[CFP],1)+6,8)," "),2))," "),"m/d/yyyy")',
                        '=IFERROR(MID(@[CFP],FIND("W/C ",@[CFP],1)+4,4)," ")', '=IFERROR(MID(@[CFP],FIND("TEMP",@[CFP],1)+4,4)," ")',
                        '= IFERROR(MID(@[CFP],FIND("G/D ",@[CFP],1)+3,6),"0") + 0', 
                        '=IF([@[Payload Number (Kg)]]="","FERRY",IF([@[Payload Number (Kg)]]=1000,"FERRY","MAX"))', 
                        '=IFERROR(MID(@[CFP],FIND("EST PLD",@[CFP],1)+7,7),"0") + 0',
                        '=IFERROR(MID(@[CFP],FIND("TRIP TIME",@[CFP],1)+9,6)," ")', 
                        '=IFERROR(MID(@[CFP],FIND("TRIP FUEL",@[CFP],1)+9,6),"0") +0', 
                        '=IFERROR(MID(@[CFP],FIND("Non-ECTRL         ",@[CFP],1)+17,8),"0") + IFERROR(MID(@[CFP],FIND("ECTRL RSO         ",@[CFP],1)+17,8),"0")',
                        '=IFERROR(MID(@[CFP],FIND("TNC-Charge        ",@[CFP],1)+17,8),"0") + 0',
                        "", copied_text, 
                        '=IFERROR(MID(@[CFP],FIND("TRIP FUEL",@[CFP],1)+9,6),"0") + IFERROR(MID(@[CFP],FIND("CONT ",@[CFP],1)+17,6),"0") + IFERROR(MID(@[CFP],FIND("TAXI              ",@[CFP],1)+17,6),"0")',
                        ]

        data.append(row_data)

    return data

# Folder path containing the text files
folder_path = r'C:\Users\Admin\Desktop\ofp' 

# Retrieve a list of text files in the folder
folder_path = folder_path + r'\*.COLLECTED'
txt_files = glob2.glob(folder_path, recursive=True)

wb = Workbook()
ws = wb.active
wb.active.title = 'Payload'


#---------------------------------------------- Sheet Configuration and Country -------------------------------------------------------------------
# Create the new sheets "Configuration" and "Country"
configuration_sheet = wb.create_sheet("Configuration", 1)
country_sheet = wb.create_sheet("Country", 2)

# Generate data for the new sheets (replace this with actual data if available)
column_headings = ["Configuration","Aircraft Registration","Aircraft used for calculation"]
config_data = [
    ["Configuration","Aircraft Registration","Aircraft used for calculation"],
    ["A320CEO","VN-A582, VN-A583, VN-A584, VN-A586, VN-A587, VN-A595","VN-A582"],
    ["A320NEO Config 1","VN-A596, VN-A598, VN-A599, VN-A226","VN-A596"],
    ["A320NEO Config 2","VN-A592, VN-A593","VN-A593"],
    ["A321CEO","VN-A585, VN-A594, VN-A597, VN-A227","VN-A594"],
    ["A321NEO Config 1","VN-A588, VN-A589, VN-A590, VN-A591 ","VN-A591"],
    ["A321NEO Config 2","VN-A222, VN-A228","VN-A222"],
    ["B787-9","VN-A818, VN-A819, VN-A829","VN-A829"]

    # Add more data as needed
]

# country_data = [
    
#     # Add more data as needed
# ]

# Populate the "Configuration" sheet with data
for row in config_data:
    configuration_sheet.append(row)

# # Populate the "Country" sheet with data
# for row in country_data:
#     country_sheet.append(row)

# Set the column width for "Configuration" sheet to 33.75
for col_num in range(1, len(column_headings) + 1):
    col_letter = get_column_letter(col_num)
    configuration_sheet.column_dimensions[col_letter].width = 33.75

# # Set the column width for "Country" sheet to 33.75
# for col_num in range(1, len(country_data[0]) + 1):
#     col_letter = get_column_letter(col_num)
#     country_sheet.column_dimensions[col_letter].width = 33.75

## Define a yellow fill for the header row
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Apply the yellow fill and border to the header row of "Configuration" sheet
for cell in configuration_sheet[1]:
    cell.fill = yellow_fill
    cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))

# # Apply the yellow fill and border to the header row of "Country" sheet
# for cell in country_sheet[1]:
#     cell.fill = yellow_fill
#     cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'), left=Side(style='medium'), right=Side(style='medium'))

# Apply borders to all cells in "Configuration" sheet (including data rows)
for row in configuration_sheet.iter_rows(min_row=1, max_row=configuration_sheet.max_row):
    for cell in row:
        cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

# # Apply borders to all cells in "Country" sheet (including data rows)
# for row in country_sheet.iter_rows(min_row=1, max_row=country_sheet.max_row):
#     for cell in row:
#         cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))



#---------------------------------------------------------------------------------------------------------------#
data = generate_data(txt_files)

# add column headings. NB. these must be strings
column_headings = ["No", "Country", "Configuration", "Departure", "Destination", "Tankering to", "Alternate",
                   "Route", "Month", "Created Date (mm/dd/yyyy)", "Wind", "Temp", "Distance (NM)", "Payload type",
                   "Payload Number (Kg)", "Flight time", "Trip fuel", "Overflight charge ($)", "TNC charge ($)",
                   "Remark", "CFP", "Fuel Burn"]
ws.append(column_headings)

for row in data:
    ws.append(row)

# Calculate the number of rows in the data, plus one for the column headings
num_rows = len(data) + 1

# Create a green fill for the header row
green_fill = PatternFill(start_color='63B052', end_color='63B052', fill_type='solid')
for cell in ws[1]:  # Apply the green fill to the entire header row
    cell.fill = green_fill

# Add a default style with white background for the rest of the table
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
for row in ws.iter_rows(min_row=2, max_row=num_rows, min_col=1, max_col=22):  # Apply white fill to all other rows
    for cell in row:
        cell.fill = white_fill

# Define a border style
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Apply border to all cells in the table
for row in ws.iter_rows(min_row=1, max_row=num_rows, min_col=1, max_col=22):
    for cell in row:
        cell.border = thin_border

# Update the "ref" parameter of the table to include all rows with data
tab = Table(displayName="Table1", ref=f"A1:V{num_rows}")

# Add a default style with striped rows and banded columns
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style

ws.add_table(tab)
# Set the width of the "No" to "Route" column to 15
ws.column_dimensions["A"].width = 15  # Assuming "Remark" column is column A (the 20th column)
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 15
ws.column_dimensions["F"].width = 15
ws.column_dimensions["G"].width = 15
ws.column_dimensions["H"].width = 15
# Set the width of the "Month" column to 15
ws.column_dimensions["I"].width = 8.29
# Set the width of the "Created Day" to "Temp" column to 15
ws.column_dimensions["J"].width = 15  # Assuming "Remark" column is column A (the 20th column)
ws.column_dimensions["K"].width = 15
ws.column_dimensions["L"].width = 15
ws.column_dimensions["M"].width = 7.57
ws.column_dimensions["N"].width = 29.86
ws.column_dimensions["O"].width = 15
ws.column_dimensions["P"].width = 15
ws.column_dimensions["Q"].width = 15
ws.column_dimensions["R"].width = 15
ws.column_dimensions["S"].width = 15
ws.column_dimensions["U"].width = 15
ws.column_dimensions["V"].width = 15
ws.column_dimensions["W"].width = 15


# Set the width of the "Remark" column to 60
ws.column_dimensions["T"].width = 60  # Assuming "Remark" column is column T (the 20th column)

wb.save(ten_file)
